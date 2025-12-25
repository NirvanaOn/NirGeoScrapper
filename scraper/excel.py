import os
import re
from openpyxl import Workbook, load_workbook

# ================= HELPERS =================

def sanitize_name(name: str) -> str:
    name = name.strip().lower()
    name = re.sub(r"[^\w\s-]", "", name)
    name = re.sub(r"\s+", "_", name)
    return name


def make_place_key(name: str, address: str) -> str:

    return f"{name.strip().lower()}|{address.strip().lower()}"


# ================= EXCEL WRITER =================

class ExcelWriter:

    def __init__(self, search_query: str, base_folder: str = "data"):

        os.makedirs(base_folder, exist_ok=True)

        safe_query = sanitize_name(search_query)
        self.path = os.path.join(base_folder, f"{safe_query}.xlsx")

        self.seen_places = set()
        self.headers = []

        if os.path.exists(self.path):
            self.wb = load_workbook(self.path)
            self.ws = self.wb.active

            if self.ws.max_row >= 1:
                self.headers = [cell.value for cell in self.ws[1] if cell.value]

            self._load_existing_places()
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Places"
            self.wb.save(self.path)

    # ================= INTERNAL =================

    def _load_existing_places(self):

        if not self.headers:
            return

        if "Name" not in self.headers or "Address" not in self.headers:
            return

        name_col = self.headers.index("Name")
        addr_col = self.headers.index("Address")

        for row in self.ws.iter_rows(min_row=2, values_only=True):
            name = row[name_col]
            address = row[addr_col]

            if name and address:
                self.seen_places.add(make_place_key(name, address))

    def _sync_headers(self, data: dict):

        new_cols = [k for k in data.keys() if k not in self.headers]
        if not new_cols:
            return

        self.headers.extend(new_cols)

        for col_index, header in enumerate(self.headers, start=1):
            self.ws.cell(row=1, column=col_index, value=header)

    # ================= PUBLIC METHODS =================

    def write_row(self, data: dict) -> bool:


        name = data.get("Name")
        address = data.get("Address")

        if not name or not address:
            return False

        key = make_place_key(name, address)

        if key in self.seen_places:
            return False

        if not self.headers:
            self.headers = list(data.keys())
            self.ws.append(self.headers)
        else:
            self._sync_headers(data)

        row = [data.get(h, "") for h in self.headers]
        self.ws.append(row)

        self.seen_places.add(key)
        self.wb.save(self.path)

        return True

    def get_row_count(self) -> int:
        return max(self.ws.max_row - 1, 0)
